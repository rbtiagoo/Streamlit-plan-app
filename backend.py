import pandas as pd
import numpy as np
import os
from datetime import datetime, timedelta
import io
import warnings
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
warnings.filterwarnings('ignore')

# Definir cores conforme solicitado
CORES = {
    'cabecalho_fill': PatternFill(start_color='8F8D8D', end_color='8F8D8D', fill_type='solid'),
    'cabecalho_font': Font(name='Aptos Narrow', size=11, color='000000', bold=True),
    'procurement_e': PatternFill(start_color='4AAFBD', end_color='4AAFBD', fill_type='solid'),
    'demand_r': PatternFill(start_color='061569', end_color='061569', fill_type='solid'),
    'demand_r_font': Font(name='Aptos Narrow', size=11, color='FFFFFF'),
    'in_stock': PatternFill(start_color='38F58A', end_color='38F58A', fill_type='solid'),
}

def criar_aba_overview(df):
    """Cria a aba Overview com as análises solicitadas"""
    
    # Criar um DataFrame vazio para a Overview
    overview_data = []
    
    # Título principal
    overview_data.append(["RELATÓRIO DE PLANEJAMENTO - VISÃO GERAL"])
    overview_data.append([])  # Linha em branco
    
    # Adicionar data de geração do relatório
    data_geracao = datetime.now().strftime("%d/%m/%Y %H:%M")
    overview_data.append([f"Relatório gerado em: {data_geracao}"])
    overview_data.append([])  # Linha em branco
    
    # Estatísticas rápidas
    overview_data.append(["ESTATÍSTICAS RÁPIDAS"])
    overview_data.append([])
    
    total_linhas = len(df)
    if 'Status' in df.columns:
        in_stock_count = len(df[df['Status'] == 'In Stock'])
    else:
        in_stock_count = 0
        
    if 'STATUS_STYPE' in df.columns:
        purrequist_count = len(df[df['STATUS_STYPE'] == 'PurRequist'])
        poconfirm_count = len(df[df['STATUS_STYPE'] == 'POConfirm'])
        pocreated_count = len(df[df['STATUS_STYPE'] == 'POCreated'])
    else:
        purrequist_count = poconfirm_count = pocreated_count = 0
    
    overview_data.append(["TOTAL DE ITENS NO RELATÓRIO:", total_linhas])
    overview_data.append(["ITENS EM ESTOQUE (IN STOCK):", in_stock_count])
    overview_data.append(["REQUISIÇÕES DE COMPRA (PURREQUIST):", purrequist_count])
    overview_data.append(["ORDENS CONFIRMADAS (POCONFIRM):", poconfirm_count])
    overview_data.append(["ORDENS CRIADAS (POCREATED):", pocreated_count])
    overview_data.append([])
    overview_data.append([])
    
    # 1. Contagem Cruzada de STATUS_STYPE por Equipment
    if 'STATUS_STYPE' in df.columns and 'Equipment' in df.columns:
        overview_data.append(["CONTAGEM CRUZADA - STATUS_STYPE POR EQUIPAMENTO"])
        overview_data.append([])
        
        contagem_cruzada = pd.crosstab(df['Equipment'], df['STATUS_STYPE'], margins=True, margins_name="TOTAL")
        
        # Adicionar cabeçalho
        header = ["EQUIPMENT"] + [str(col).upper() for col in contagem_cruzada.columns]
        overview_data.append(header)
        
        # Adicionar dados
        for equipamento in contagem_cruzada.index:
            linha = [equipamento]
            for stype in contagem_cruzada.columns:
                linha.append(contagem_cruzada.loc[equipamento, stype])
            overview_data.append(linha)
        
        overview_data.append([])
        overview_data.append([])
    
    # 2. Itens com atraso Top 15 - VERSÃO SIMPLIFICADA
    overview_data.append(["ITENS COM ATRASO - TOP 15"])
    overview_data.append(["MATERIAL_NO", "MATERIAL_DESCRIPTION", "FLOAT(TODAY-OPENING)"])
    
    if all(col in df.columns for col in ['MATERIAL_NO', 'MATERIAL_DESCRIPTION', 'Float(Today-Opening)']):
        # Criar cópia do DataFrame
        df_atraso = df.copy()
        
        # Converter a coluna Float(Today-Opening) para numérico
        df_atraso['Float(Today-Opening)'] = pd.to_numeric(df_atraso['Float(Today-Opening)'], errors='coerce')
        
        # Remover linhas com valores NaN na coluna Float(Today-Opening)
        df_atraso = df_atraso.dropna(subset=['Float(Today-Opening)'])
        
        # Ordenar do menor para o maior (incluindo valores negativos)
        df_atraso = df_atraso.sort_values('Float(Today-Opening)', ascending=True)
        
        # Pegar os 15 primeiros itens
        top_atraso = df_atraso.head(15)[['MATERIAL_NO', 'MATERIAL_DESCRIPTION', 'Float(Today-Opening)']]
        
        # Adicionar à overview
        for _, row in top_atraso.iterrows():
            overview_data.append([
                row['MATERIAL_NO'],
                row['MATERIAL_DESCRIPTION'],
                row['Float(Today-Opening)']
            ])
    else:
        overview_data.append(["COLUNAS NECESSÁRIAS PARA ANÁLISE DE ATRASO NÃO ENCONTRADAS"])
    
    overview_data.append([])
    overview_data.append([])
    
    # 3. Itens de Montagem (E) próximo período
    overview_data.append(["ITENS DE MONTAGEM (E) - PRÓXIMOS 3 MESES"])
    overview_data.append(["MATERIAL_NO", "MATERIAL_DESCRIPTION", "OPENING_DATE"])
    
    if all(col in df.columns for col in ['PROCUREMENT_KEY', 'MATERIAL_NO', 'MATERIAL_DESCRIPTION', 'OPENING_DATE']):
        df_montagem = df[df['PROCUREMENT_KEY'] == 'E'].copy()
        
        if not df_montagem.empty:
            # Converter OPENING_DATE para datetime
            df_montagem['OPENING_DATE_DT'] = pd.to_datetime(df_montagem['OPENING_DATE'], format='%d/%m/%Y', errors='coerce')
            
            # Filtrar para próximos 3 meses
            hoje = datetime.now()
            tres_meses = hoje + timedelta(days=90)
            
            df_montagem_filtrado = df_montagem[
                (df_montagem['OPENING_DATE_DT'] >= hoje) & 
                (df_montagem['OPENING_DATE_DT'] <= tres_meses)
            ].copy()
            
            if not df_montagem_filtrado.empty:
                # Ordenar da menor data para a maior
                df_montagem_filtrado = df_montagem_filtrado.sort_values('OPENING_DATE_DT', ascending=True)
                
                for _, row in df_montagem_filtrado.iterrows():
                    overview_data.append([
                        row['MATERIAL_NO'],
                        row['MATERIAL_DESCRIPTION'],
                        row['OPENING_DATE']
                    ])
            else:
                overview_data.append(["NÃO HÁ MONTAGENS PREVISTAS PARA O PERÍODO DE 90 DIAS"])
                overview_data.append([])
        else:
            overview_data.append(["NÃO HÁ ITENS DE MONTAGEM (PROCUREMENT_KEY = 'E') NO ARQUIVO"])
            overview_data.append([])
    else:
        overview_data.append(["COLUNAS NECESSÁRIAS PARA ANÁLISE DE MONTAGEM NÃO ENCONTRADAS"])
        overview_data.append([])
    
    # Converter para DataFrame
    df_overview = pd.DataFrame(overview_data)
    
    return df_overview

def aplicar_fonte_global(worksheet):
    """Aplica fonte Aptos Narrow tamanho 11 para todas as células da worksheet"""
    for row in worksheet.iter_rows():
        for cell in row:
            # Manter o negrito existente se já estiver aplicado, mas alterar a fonte e tamanho
            current_bold = cell.font.bold if cell.font else False
            cell.font = Font(name='Aptos Narrow', size=11, bold=current_bold)

def aplicar_formato_overview(worksheet):
    """Aplica formatação específica para a aba Overview - apenas negrito em títulos e cabeçalhos"""
    for row in worksheet.iter_rows():
        for cell in row:
            # Formatar títulos principais e cabeçalhos de tabela em negrito
            if cell.value and any(titulo in str(cell.value) for titulo in [
                "RELATÓRIO DE PLANEJAMENTO - VISÃO GERAL",
                "Relatório gerado em:",
                "ESTATÍSTICAS RÁPIDAS",
                "CONTAGEM CRUZADA - STATUS_STYPE POR EQUIPAMENTO", 
                "ITENS COM ATRASO - TOP 15",
                "ITENS DE MONTAGEM (E) - PRÓXIMOS 3 MESES"
            ]):
                cell.font = Font(name='Aptos Narrow', size=11, bold=True)
                
            # Formatar cabeçalhos de colunas em negrito
            elif cell.value and str(cell.value) in ["EQUIPMENT", "MATERIAL_NO", "MATERIAL_DESCRIPTION", 
                                                   "FLOAT(TODAY-OPENING)", "OPENING_DATE", "TOTAL"]:
                cell.font = Font(name='Aptos Narrow', size=11, bold=True)
                
            # Formatar estatísticas rápidas em negrito
            elif cell.value and any(stat in str(cell.value) for stat in [
                "TOTAL DE ITENS NO RELATÓRIO:",
                "ITENS EM ESTOQUE (IN STOCK):",
                "REQUISIÇÕES DE COMPRA (PURREQUIST):",
                "ORDENS CONFIRMADAS (POCONFIRM):",
                "ORDENS CRIADAS (POCREATED):"
            ]):
                cell.font = Font(name='Aptos Narrow', size=11, bold=True)

def aplicar_formato_condicional(worksheet, df):
    """Aplica formatação condicional baseada nas regras especificadas"""
    
    # Obter índices das colunas relevantes
    header = [cell.value for cell in worksheet[1]]
    
    # Encontrar índices das colunas que precisamos
    col_procurement_key = None
    col_demand = None
    col_status = None
    
    for idx, cell in enumerate(worksheet[1]):
        if cell.value == "PROCUREMENT_KEY":
            col_procurement_key = idx
        elif cell.value == "DEMAND":
            col_demand = idx
        elif cell.value == "Status":
            col_status = idx
    
    # Formatar cabeçalho
    for cell in worksheet[1]:
        cell.fill = CORES['cabecalho_fill']
        cell.font = CORES['cabecalho_font']
    
    # Aplicar formatação condicional às linhas de dados
    for row in worksheet.iter_rows(min_row=2):
        # Verificar condições na ordem de prioridade
        
        # 1. Status = "In Stock" (maior prioridade)
        if col_status is not None:
            status_cell = row[col_status]
            if status_cell.value == "In Stock":
                for cell in row:
                    cell.fill = CORES['in_stock']
                continue  # Pula para próxima linha
        
        # 2. DEMAND começa com "R"
        if col_demand is not None:
            demand_cell = row[col_demand]
            if demand_cell.value and str(demand_cell.value).startswith('R'):
                for cell in row:
                    cell.fill = CORES['demand_r']
                    cell.font = CORES['demand_r_font']
                continue  # Pula para próxima linha
        
        # 3. PROCUREMENT_KEY = "E"
        if col_procurement_key is not None:
            procurement_cell = row[col_procurement_key]
            if procurement_cell.value == "E":
                for cell in row:
                    cell.fill = CORES['procurement_e']

def aplicar_formato_excel(writer, df_dict):
    """Aplica formatação colorida ao arquivo Excel baseada nas condições especificadas"""
    workbook = writer.book
    
    for sheet_name in df_dict:
        worksheet = workbook[sheet_name]
        
        # Aplicar fonte Aptos Narrow tamanho 11 para todas as células
        aplicar_fonte_global(worksheet)
        
        if sheet_name != 'Overview':
            # Aplicar formatação específica para outras abas
            aplicar_formato_condicional(worksheet, df_dict[sheet_name])
        else:
            # Aplicar formatação específica para Overview (apenas negrito em títulos)
            aplicar_formato_overview(worksheet)
        
        # Ajustar largura das colunas com limite de 270 pixels (35 caracteres)
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        max_length = max(max_length, cell_length)
                except:
                    pass
            adjusted_width = min(max_length + 2, 35)
            worksheet.column_dimensions[column_letter].width = adjusted_width

def processar_arquivo(uploaded_file):
    """Função principal para processar o arquivo"""
    
    # Lista de colunas de data para converter
    colunas_data_especificas = [
        "Data de necessidade", "Sup Date or Log Date", "OPENING_DATE", 
        "Opening_Calculada", "PURCHASING_DOC_DATE", "DELIVERY_DATE", 
        "SUPPLY_DATE", "REQUIRED_DATE", "Data_Atual"
    ]
    
    try:
        # Carregar o arquivo
        df = pd.read_excel(uploaded_file)
        
        # Converter colunas de data
        for col in colunas_data_especificas:
            if col in df.columns:
                try:
                    df[col] = pd.to_datetime(df[col], errors='coerce')
                    df[col] = df[col].dt.strftime('%d/%m/%Y')
                except:
                    pass
        
        # Ordenar por Counter se existir
        if 'Counter' in df.columns:
            df['Counter'] = pd.to_numeric(df['Counter'], errors='coerce')
            df = df.sort_values(by='Counter', ascending=True)
        
        # Criar aba RM
        colunas_rm = [
            "RESPONSIBLE", "Escopo", "Equipment", "Prazo", "Status", "XP_STATUS",
            "MATERIAL_NO", "MATERIAL_DESCRIPTION", "STATUS_STYPE", "PRREQRELSTAT",
            "PO", "DOC PGR", "BUYER_NAME", "OPENING_DATE", "Float(Today-Opening)", "DELIVERY_DATE",
            "SUPPLY_DATE", "REQUIRED_DATE", "FLOAT", "PLANNED_DELIVERY_TIME_MM",
            "IN_HOUSE_PROD_TIME", "GRP_TIME_MM", "WBS", "EXCEPTION_MESSAGE", "Comentários SAP Planner"
        ]
        
        colunas_disponiveis_rm = [col for col in colunas_rm if col in df.columns]
        df_rm = pd.DataFrame()
        if 'STATUS_STYPE' in df.columns:
            df_rm = df[df['STATUS_STYPE'] == 'PurRequist'][colunas_disponiveis_rm].copy()
        
        # Criar aba PO
        colunas_po = [
            "RESPONSIBLE", "Escopo", "Equipment", "Prazo", "Status", "XP_STATUS",
            "MATERIAL_NO", "MATERIAL_DESCRIPTION", "STATUS_STYPE", "PO", "DOC PGR", "BUYER_NAME", 
            "DELIVERY_DATE", "SUPPLY_DATE", "REQUIRED_DATE", "FLOAT", "VENDOR_NAME", "WBS", 
            "EXCEPTION_MESSAGE", "Comentários SAP Planner", "Comentario LogPlan"
        ]
        
        colunas_disponiveis_po = [col for col in colunas_po if col in df.columns]
        df_po = pd.DataFrame()
        if 'STATUS_STYPE' in df.columns:
            df_po = df[df['STATUS_STYPE'].isin(['POConfirm', 'POCreated'])][colunas_disponiveis_po].copy()

        # Criar aba Stock
        colunas_stock = [
            "RESPONSIBLE", "Equipment", "Prazo", "Status", "XP_STATUS",
            "DEMAND", "MATERIAL_NO", "MATERIAL_DESCRIPTION", "STATUS_STYPE",
            "PO", "REQUIRED_DATE", "WBS", "EXCEPTION_MESSAGE", "QN_NUMBER", 
            "QN_COORDINATOR", "TYPE_310_", "TYPE_321_", "TYPE_999_", "ECN1", 
            "ABRG_", "Comentários SAP Planner"
        ]
        
        colunas_disponiveis_stock = [col for col in colunas_stock if col in df.columns]
        df_stock = pd.DataFrame()
        if 'Status' in df.columns:
            df_stock = df[df['Status'] == 'In Stock'][colunas_disponiveis_stock].copy()

        # Criar aba Overview
        df_overview = criar_aba_overview(df)
        
        # Criar arquivo Excel em memória
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Salvar as abas na ordem desejada
            df_overview.to_excel(writer, sheet_name='Overview', index=False, header=False)
            df.to_excel(writer, sheet_name='Relatório BI', index=False)
            df_rm.to_excel(writer, sheet_name='RM', index=False)
            df_po.to_excel(writer, sheet_name='PO', index=False)
            df_stock.to_excel(writer, sheet_name='Stock', index=False)
            
            # Aplicar formatação
            df_dict = {
                'Overview': df_overview,
                'Relatório BI': df,
                'RM': df_rm,
                'PO': df_po,
                'Stock': df_stock
            }
            aplicar_formato_excel(writer, df_dict)
            
            # Formatação básica (filtros e congelamento)
            for sheet_name in writer.sheets:
                worksheet = writer.sheets[sheet_name]
                if sheet_name != 'Overview':  # Não aplicar filtro na Overview
                    worksheet.auto_filter.ref = worksheet.dimensions
                    worksheet.freeze_panes = 'A2'
        
        output.seek(0)
        return output, len(df), len(df_rm), len(df_po), len(df_stock)
        
    except Exception as e:
        raise Exception(f"Erro no processamento: {str(e)}")